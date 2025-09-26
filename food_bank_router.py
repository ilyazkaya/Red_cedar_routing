from __future__ import annotations

import csv
import math
from datetime import datetime
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import requests
import xml.etree.ElementTree as ET

try:
    from ortools.graph import pywrapgraph
except ImportError:
    pywrapgraph = None

NOMINATIM_EMAIL = "ilyakapral@gmail.com"
VICTORIA_LAT = 48.4284
VICTORIA_LON = -123.3656
VICTORIA_BOUNDING_BOX = {
    "min_lat": 48.35,
    "max_lat": 48.60,
    "min_lon": -123.50,
    "max_lon": -123.20,
}

VICTORIA_VIEWBOX = f"{VICTORIA_BOUNDING_BOX['min_lon']},{VICTORIA_BOUNDING_BOX['min_lat']},{VICTORIA_BOUNDING_BOX['max_lon']},{VICTORIA_BOUNDING_BOX['max_lat']}"

MAX_ORDERS_PER_ROUTE = 20
OUTPUT_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"

KML_ROUTE_COLORS = [
    "ff0000ff",  # red
    "ff00ff00",  # green
    "ffff0000",  # blue
    "ff00ffff",  # yellow
    "ffff00ff",  # magenta
    "ffffff00",  # cyan
    "ff7f00ff",  # orange
    "ff007fff",  # purple
]
MCMF_DISTANCE_COST_MULTIPLIER = 1000

@dataclass
class Route:
    name: str  # human readable route name
    lat: Optional[float] = None  # optional route center latitude
    lon: Optional[float] = None  # optional route center longitude

    @staticmethod
    def load_from_csv(csv_path: str | Path) -> Dict[str, "Route"]:
        routes: Dict[str, Route] = {}  # keyed by route name for quick lookup
        path = Path(csv_path)
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                name = (row.get("region") or "").strip()  # region label from file
                if not name:
                    continue
                lat = Route._parse_float(row.get("lat"))  # parsed latitude value
                lon = Route._parse_float(row.get("lon"))  # parsed longitude value
                routes[name] = Route(name=name, lat=lat, lon=lon)
        return routes

    @staticmethod
    def _parse_float(value: Optional[str]) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None


@dataclass
class Order:
    id: Optional[int] = None  # unique order identifier when available
    name: str = ""  # customer name
    phone: str = ""  # customer phone number
    email: str = ""  # customer email address
    address1: str = ""  # primary street address line
    address2: str = ""  # secondary address details (suite, notes)
    delivery_instructions: str = ""  # additional delivery directions
    lat: Optional[float] = None  # address latitude
    lon: Optional[float] = None  # address longitude
    route: Optional[Route] = None  # assigned route metadata
    route_position: Optional[int] = None  # sequence slot inside the route

    def address_cell(self) -> str:
        if self.address2:
            return f"{self.address1}\n{self.address2}"
        return self.address1


class OrderBook:
    WORKING_HEADERS = [
        "id",
        "name",
        "phone",
        "email",
        "address1",
        "address2",
        "delivery_instructions",
        "lat",
        "lon",
        "route",
        "route_position",
    ]

    def __init__(self, orders: Optional[Iterable[Order]] = None):
        self.orders: List[Order] = list(orders or [])  # normalized list of orders

    @classmethod
    def from_external_csv(
        cls,
        csv_path: str | Path,
        routes: Optional[Dict[str, Route]] = None,
    ) -> "OrderBook":
        path = Path(csv_path)
        orders: List[Order] = []  # collected orders parsed from file
        position_map: Dict[str, int] = {}  # tracks shared route positions per address
        position_counter = 0  # next route position to assign

        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for idx, row in enumerate(reader, start=1):
                address_raw = (row.get("Address") or "").strip()
                address_parts = [piece.strip() for piece in address_raw.splitlines() if piece.strip()]  # split address into lines
                address1 = address_parts[0] if address_parts else ""
                address2 = address_parts[1] if len(address_parts) > 1 else ""

                key = address1.lower()  # normalized address key used for grouping
                if key and key in position_map:
                    route_position = position_map[key]
                else:
                    position_counter += 1
                    route_position = position_counter
                    if key:
                        position_map[key] = route_position

                route_name = (row.get("Route") or "").strip()
                route_obj = cls._resolve_route(route_name, routes)

                orders.append(
                    Order(
                        id=idx,
                        name=(row.get("Name") or "").strip(),
                        phone=(row.get("Phone") or "").strip(),
                        email=(row.get("Email") or "").strip(),
                        address1=address1,
                        address2=address2,
                        delivery_instructions=(row.get("Delivery Instructions") or "").strip(),
                        route=route_obj,
                        route_position=route_position,
                    )
                )
        return cls(orders)

    @classmethod
    def from_working_csv(
        cls,
        csv_path: str | Path,
        routes: Optional[Dict[str, Route]] = None,
    ) -> "OrderBook":
        path = Path(csv_path)
        orders: List[Order] = []  # collected orders from working file
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not any(row.values()):
                    continue
                route_name = (row.get("route") or "").strip()
                route_obj = cls._resolve_route(route_name, routes)

                orders.append(
                    Order(
                        id=cls._parse_int(row.get("id")),
                        name=(row.get("name") or "").strip(),
                        phone=(row.get("phone") or "").strip(),
                        email=(row.get("email") or "").strip(),
                        address1=(row.get("address1") or "").strip(),
                        address2=(row.get("address2") or "").strip(),
                        delivery_instructions=(row.get("delivery_instructions") or "").strip(),
                        lat=Route._parse_float(row.get("lat")),
                        lon=Route._parse_float(row.get("lon")),
                        route=route_obj,
                        route_position=cls._parse_int(row.get("route_position")),
                    )
                )
        return cls(orders)

    def to_working_csv(self, csv_path: str | Path) -> None:
        path = Path(csv_path)
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=self.WORKING_HEADERS)
            writer.writeheader()
            for order in self.orders:
                writer.writerow(
                    {
                        "id": order.id if order.id is not None else "",
                        "name": order.name,
                        "phone": order.phone,
                        "email": order.email,
                        "address1": order.address1,
                        "address2": order.address2,
                        "delivery_instructions": order.delivery_instructions,
                        "lat": _format_float(order.lat),
                        "lon": _format_float(order.lon),
                        "route": order.route.name if order.route else "",
                        "route_position": order.route_position if order.route_position is not None else "",
                    }
                )

    


    def to_simple_xlsx(self, xlsx_path: str | Path) -> None:
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - dependency hint
            raise RuntimeError("openpyxl is required for exporting xlsx files") from exc

        path = Path(xlsx_path)
        workbook = Workbook()
        workbook.remove(workbook.active)

        headers = ["Address", "Name", "Phone", "Email", "Delivery Instructions"]  # column structure for each sheet
        route_names = self._ordered_route_names()  # ordered list of route sheet names
        used_titles: Dict[str, int] = {}  # tracks sheet name collisions

        summary_sheet = workbook.create_sheet(title="Route Summary")
        summary_sheet.append(["Route", "Orders", "Stops"])

        for route_name in route_names:
            orders_in_route = self._orders_for_route(route_name)
            sheet_title = self._make_sheet_title(route_name, used_titles)
            sheet = workbook.create_sheet(title=sheet_title)
            sheet.append(headers)

            unique_stops = {
                order.address1.strip().lower()
                for order in orders_in_route
                if order.address1.strip()
            }
            summary_sheet.append([route_name, len(orders_in_route), len(unique_stops)])

            for order in orders_in_route:
                sheet.append(
                    [
                        order.address_cell(),
                        order.name,
                        order.phone,
                        order.email,
                        order.delivery_instructions,
                    ]
                )

        workbook.save(path)

    def to_kml(self, kml_path: str | Path, connect_points: bool = True) -> None:
        missing = [order for order in self.orders if order.lat is None or order.lon is None]  # orders lacking coordinates
        if missing:
            identifiers = ", ".join(str(order.id or order.address1) for order in missing)
            raise ValueError(f"Cannot export KML, missing coordinates for: {identifiers}")

        path = Path(kml_path)
        kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
        document = ET.SubElement(kml, "Document")

        route_names = self._ordered_route_names()
        style_ids: Dict[str, str] = {}
        for idx, route_name in enumerate(route_names):
            color = KML_ROUTE_COLORS[idx % len(KML_ROUTE_COLORS)]
            style_id = f"route-style-{idx}"
            style = ET.SubElement(document, "Style", id=style_id)

            icon_style = ET.SubElement(style, "IconStyle")
            ET.SubElement(icon_style, "color").text = color
            ET.SubElement(icon_style, "scale").text = "1.1"
            icon = ET.SubElement(icon_style, "Icon")
            ET.SubElement(icon, "href").text = "http://maps.google.com/mapfiles/kml/paddle/wht-blank.png"

            line_style = ET.SubElement(style, "LineStyle")
            ET.SubElement(line_style, "color").text = color
            ET.SubElement(line_style, "width").text = "2"

            style_ids[route_name] = style_id

        for route_name in route_names:
            folder = ET.SubElement(document, "Folder")
            ET.SubElement(folder, "name").text = route_name

            coordinates: List[Tuple[Optional[int], str]] = []  # coordinates paired with their route position
            seen_positions: set[int] = set()  # prevents duplicate stops at same position

            for order in self._orders_for_route(route_name):
                if order.route_position is not None and order.route_position in seen_positions:
                    continue
                if order.route_position is not None:
                    seen_positions.add(order.route_position)

                placemark = ET.SubElement(folder, "Placemark")
                style_id = style_ids.get(route_name)
                if style_id:
                    ET.SubElement(placemark, "styleUrl").text = f"#{style_id}"
                label = "" if order.route_position is None else str(order.route_position)
                ET.SubElement(placemark, "name").text = label
                ET.SubElement(placemark, "description").text = _address_with_notes(order)

                point = ET.SubElement(placemark, "Point")
                coord = f"{_format_float(order.lon)},{_format_float(order.lat)},0"  # KML expects lon,lat,altitude
                ET.SubElement(point, "coordinates").text = coord
                coordinates.append((order.route_position, coord))

            if connect_points:
                ordered = [
                    coord
                    for _, coord in sorted(
                        coordinates,
                        key=lambda item: (
                            item[0] is None,
                            item[0] if item[0] is not None else float("inf"),
                        ),
                    )
                ]
                if len(ordered) >= 2:
                    line = ET.SubElement(folder, "Placemark")
                    style_id = style_ids.get(route_name)
                    if style_id:
                        ET.SubElement(line, "styleUrl").text = f"#{style_id}"
                    ET.SubElement(line, "name").text = "Route"
                    line_string = ET.SubElement(line, "LineString")
                    ET.SubElement(line_string, "coordinates").text = " ".join(ordered)

        tree = ET.ElementTree(kml)
        tree.write(path, encoding="utf-8", xml_declaration=True)

    def assign_routes(
        self,
        routes: Dict[str, Route],
        max_orders_per_route: int = MAX_ORDERS_PER_ROUTE,
    ) -> None:
        if pywrapgraph is None:
            raise RuntimeError(
                "OR-Tools is required for min-cost flow route assignment. Install 'ortools' to continue."
            )
        if not routes:
            raise ValueError("No routes provided for assignment.")

        centers: Dict[str, Route] = {}
        for name, route in routes.items():
            if route.lat is None or route.lon is None:
                raise ValueError(f"Route '{name}' is missing coordinates.")
            centers[name] = route

        route_names = list(centers.keys())
        route_count = len(route_names)
        if route_count == 0:
            raise ValueError("No usable routes found.")
        route_index: Dict[str, int] = {name: idx for idx, name in enumerate(route_names)}

        stops_map: Dict[str, Dict[str, object]] = {}
        for order in self.orders:
            address = order.address1.strip()
            if not address:
                raise ValueError(f"Order {order.id or order.name or 'unknown'} has empty address1.")
            if order.lat is None or order.lon is None:
                raise ValueError(
                    f"Order {order.id or address} lacks coordinates; geocode before assigning routes."
                )
            key = address.lower()
            record = stops_map.setdefault(
                key,
                {"address": address, "orders": [], "lat": order.lat, "lon": order.lon},
            )
            record["orders"].append(order)
            if record.get("lat") is None:
                record["lat"] = order.lat
            if record.get("lon") is None:
                record["lon"] = order.lon

        stops: List[Dict[str, object]] = []
        for record in stops_map.values():
            orders = record["orders"]
            lat = record.get("lat")
            lon = record.get("lon")
            if lat is None or lon is None:
                raise ValueError(f"Address '{record['address']}' lacks coordinates.")
            order_count = len(orders)
            if order_count > max_orders_per_route:
                raise ValueError(
                    f"Address '{record['address']}' has {order_count} orders, which exceeds the per-route cap of {max_orders_per_route}."
                )
            options: List[Tuple[int, float]] = []
            costs_by_route: Dict[int, int] = {}
            for name in route_names:
                idx = route_index[name]
                center = centers[name]
                assert center.lat is not None and center.lon is not None
                distance = _distance_km(lat, lon, center.lat, center.lon)
                cost = int(round(distance * MCMF_DISTANCE_COST_MULTIPLIER))
                options.append((idx, distance))
                costs_by_route[idx] = cost
            options.sort(key=lambda item: item[1])
            stops.append(
                {
                    "address": record["address"],
                    "orders": orders,
                    "lat": lat,
                    "lon": lon,
                    "order_count": order_count,
                    "options": options,
                    "costs": costs_by_route,
                }
            )

        total_orders = sum(stop["order_count"] for stop in stops)
        total_stops = len(stops)
        capacity_total = max_orders_per_route * route_count
        if total_orders > capacity_total:
            raise ValueError(
                f"Not enough route capacity ({capacity_total}) for {total_orders} orders. Increase MAX_ORDERS_PER_ROUTE or add routes."
            )
        if total_stops == 0:
            print("No stops available for route assignment; skipping.")
            return

        def _balanced_stop_capacities(stop_total: int, route_total: int) -> List[int]:
            base = stop_total // route_total
            remainder = stop_total % route_total
            return [base + (1 if idx < remainder else 0) for idx in range(route_total)]

        stop_capacities = _balanced_stop_capacities(total_stops, route_count)
        print(
            f"Running min-cost flow assignment for {total_stops} stops across {route_count} routes."
        )

        def _solve_flow(disabled_pairs: set[Tuple[int, int]]) -> List[int]:
            for stop_idx, stop in enumerate(stops):
                has_arc = any(
                    stop_capacities[route_idx] > 0 and (stop_idx, route_idx) not in disabled_pairs
                    for route_idx, _ in stop["options"]
                )
                if not has_arc:
                    raise ValueError(
                        f"No available routes remain for address '{stop['address']}'."
                    )

            solver = pywrapgraph.SimpleMinCostFlow()
            source = 0
            stop_offset = 1
            route_offset = stop_offset + total_stops
            sink = route_offset + route_count

            for stop_idx in range(total_stops):
                solver.AddArcWithCapacityAndUnitCost(source, stop_offset + stop_idx, 1, 0)

            for stop_idx, stop in enumerate(stops):
                for route_idx, _ in stop["options"]:
                    if stop_capacities[route_idx] <= 0:
                        continue
                    if (stop_idx, route_idx) in disabled_pairs:
                        continue
                    solver.AddArcWithCapacityAndUnitCost(
                        stop_offset + stop_idx,
                        route_offset + route_idx,
                        1,
                        stop["costs"][route_idx],
                    )

            for route_idx, capacity in enumerate(stop_capacities):
                if capacity <= 0:
                    continue
                solver.AddArcWithCapacityAndUnitCost(route_offset + route_idx, sink, capacity, 0)

            solver.SetNodeSupply(source, total_stops)
            solver.SetNodeSupply(sink, -total_stops)
            for node in range(stop_offset, stop_offset + total_stops):
                solver.SetNodeSupply(node, 0)
            for node in range(route_offset, route_offset + route_count):
                solver.SetNodeSupply(node, 0)

            status = solver.Solve()
            if status != solver.OPTIMAL:
                raise ValueError(f"Min-cost flow solver failed with status {status}.")

            assignments = [-1] * total_stops
            for arc_idx in range(solver.NumArcs()):
                tail = solver.Tail(arc_idx)
                head = solver.Head(arc_idx)
                if stop_offset <= tail < route_offset and route_offset <= head < sink:
                    if solver.Flow(arc_idx) > 0:
                        stop_idx = tail - stop_offset
                        route_idx = head - route_offset
                        assignments[stop_idx] = route_idx

            if any(route_idx == -1 for route_idx in assignments):
                raise ValueError("Solver returned incomplete route assignments.")
            return assignments

        banned_pairs: set[Tuple[int, int]] = set()
        max_iterations = max(1, len(stops) * route_count * 2)
        final_assignments: List[int] = []
        route_stop_totals: Dict[str, int] = {}
        route_order_totals: Dict[str, int] = {}

        for iteration in range(max_iterations):
            assignments = _solve_flow(banned_pairs)
            route_stop_totals = {name: 0 for name in route_names}
            route_order_totals = {name: 0 for name in route_names}
            for stop_idx, route_idx in enumerate(assignments):
                route_name = route_names[route_idx]
                route_stop_totals[route_name] += 1
                route_order_totals[route_name] += stops[stop_idx]["order_count"]

            overflow = [name for name, total in route_order_totals.items() if total > max_orders_per_route]
            if not overflow:
                final_assignments = assignments
                print("Min-cost flow assignment complete.")
                break

            print(
                "Order cap exceeded for routes: " + ", ".join(overflow) + ". Attempting rebalance."
            )
            move_made = False
            for route_name in overflow:
                route_idx = route_index[route_name]
                candidates: List[Tuple[float, int, int]] = []
                for stop_idx, assigned_idx in enumerate(assignments):
                    if assigned_idx != route_idx:
                        continue
                    current_cost = stops[stop_idx]["costs"][route_idx]
                    for alt_route_idx, _ in stops[stop_idx]["options"]:
                        if alt_route_idx == route_idx:
                            continue
                        if stop_capacities[alt_route_idx] <= 0:
                            continue
                        if (stop_idx, alt_route_idx) in banned_pairs:
                            continue
                        alt_route_name = route_names[alt_route_idx]
                        projected_orders = (
                            route_order_totals[alt_route_name] + stops[stop_idx]["order_count"]
                        )
                        if projected_orders > max_orders_per_route:
                            continue
                        delta_cost = stops[stop_idx]["costs"][alt_route_idx] - current_cost
                        candidates.append((delta_cost, -stops[stop_idx]["order_count"], stop_idx))
                        break
                if candidates:
                    candidates.sort()
                    _, _, chosen_stop_idx = candidates[0]
                    banned_pairs.add((chosen_stop_idx, route_idx))
                    print(
                        f"Rebalancing: preventing stop '{stops[chosen_stop_idx]['address']}' from route '{route_name}'."
                    )
                    move_made = True
                    break

            if not move_made:
                raise ValueError(
                    "Unable to satisfy per-route order cap with available routes. Increase MAX_ORDERS_PER_ROUTE or add routes."
                )
        else:
            raise ValueError("Exceeded max attempts while enforcing order caps.")

        if not final_assignments:
            raise ValueError("Route assignment failed to converge.")

        for stop_idx, route_idx in enumerate(final_assignments):
            route_name = route_names[route_idx]
            route_obj = centers[route_name]
            for order in stops[stop_idx]["orders"]:
                order.route = route_obj
                order.route_position = None

        for route_name in route_names:
            print(
                f"Route {route_name}: {route_stop_totals[route_name]} stops, {route_order_totals[route_name]} orders."
            )
    def _orders_for_route(self, route_name: str) -> List[Order]:
        orders = [order for order in self.orders if self._route_name(order) == route_name]
        return sorted(
            orders,
            key=lambda order: (
                order.route_position if order.route_position is not None else float("inf"),
                order.id if order.id is not None else 0,
            ),
        )

    def _ordered_route_names(self) -> List[str]:
        names: List[str] = []  # preserves first-seen order of route names
        seen = set()
        for order in self.orders:
            name = self._route_name(order)
            if name not in seen:
                seen.add(name)
                names.append(name)
        return names

    @staticmethod
    def _route_name(order: Order) -> str:
        return order.route.name if order.route else "Unassigned"

    @staticmethod
    def _resolve_route(route_name: str, routes: Optional[Dict[str, Route]]) -> Optional[Route]:
        if not route_name:
            return None
        if routes and route_name in routes:
            return routes[route_name]
        return Route(name=route_name)

    @staticmethod
    def _parse_int(value: Optional[str]) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(value)
        except ValueError:
            return None

    @staticmethod
    def _make_sheet_title(name: str, counts: Dict[str, int]) -> str:
        base = (name or "Unassigned")[:31]  # Excel imposes a 31 character limit
        count = counts.get(base, 0)  # how many times we have used this base name
        title = base if count == 0 else OrderBook._with_suffix(base, count)
        while title in counts:
            count += 1
            title = OrderBook._with_suffix(base, count)
        counts[base] = count
        counts[title] = 0
        return title

    @staticmethod
    def _with_suffix(base: str, count: int) -> str:
        if count <= 0:
            return base
        suffix = f"_{count}"
        limit = max(0, 31 - len(suffix))  # ensure suffix does not exceed Excel limit
        trimmed = base[:limit] if limit else ""
        result = f"{trimmed}{suffix}"
        return result[:31]

def geocode_orders(orders: Iterable[Order], cache_path: Path, delay_seconds: float = 1.0) -> None:
    cache_path = Path(cache_path)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if not cache_path.exists():
        with cache_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["address", "lat", "lon"])

    cache, canonical, entry_count, has_header = _load_address_cache(cache_path)
    if (entry_count and entry_count != len(canonical)) or (entry_count and not has_header):
        _rewrite_cache_file(cache_path, canonical)
        cache, canonical, entry_count, has_header = _load_address_cache(cache_path)

    session = requests.Session()
    session.headers.update({"User-Agent": f"FoodBankRouter/1.0 ({NOMINATIM_EMAIL})"})

    pending = [order for order in orders if order.lat is None or order.lon is None]
    if not pending:
        return

    last_request = 0.0
    for order in pending:
        key = _cache_key(order)
        if not key:
            raise ValueError(f"Cannot geocode order {order.id or order.name}, address is empty.")

        if key in cache:
            print(f'Found {order.address1} in cache! Skipping')
            order.lat, order.lon = cache[key]
            continue

        wait_for = delay_seconds - (time.time() - last_request)
        if wait_for > 0:
            time.sleep(wait_for)

        print(f"requesting Nominatim geocode for {order.address1}")
        try:
            response = session.get(
                "https://nominatim.openstreetmap.org/search",
                params={
                    "q": order.address1.strip(),
                    "format": "json",
                    "addressdetails": 0,
                    "limit": 5,
                    "countrycodes": "ca",
                    "viewbox": VICTORIA_VIEWBOX,
                    "bounded": 1,
                    "email": NOMINATIM_EMAIL,
                },
                timeout=20,
            )
            response.raise_for_status()
        except requests.RequestException as exc:
            raise RuntimeError(f"Nominatim request failed for '{_address_with_notes(order, ' | ')}': {exc}") from exc

        last_request = time.time()
        results = response.json()
        if not results:
            raise ValueError(f"Nominatim could not find address: {_address_with_notes(order, ' | ')}")

        print(results)
        ordered_candidates = sorted(
            results,
            key=lambda item: _distance_km(float(item["lat"]), float(item["lon"]), VICTORIA_LAT, VICTORIA_LON),
        )

        best = None
        for candidate in ordered_candidates:
            candidate_lat = float(candidate["lat"])
            candidate_lon = float(candidate["lon"])
            if _within_victoria(candidate_lat, candidate_lon):
                best = candidate
                break

        if best is None:
            sample = ordered_candidates[0]
            sample_lat = float(sample["lat"])
            sample_lon = float(sample["lon"])
            raise ValueError(
                f"No geocode results within Victoria bounds for '{_address_with_notes(order, ' | ')}'. Closest match lat={sample_lat}, lon={sample_lon}"
            )

        lat = float(best["lat"])
        lon = float(best["lon"])

        order.lat = lat
        order.lon = lon
        canonical_address = order.address1.strip()
        _register_cache_entry(cache, canonical, canonical_address, lat, lon)
        _append_cache_entry(cache_path, canonical_address, lat, lon)



def _cache_key(order: Order) -> str:
    return order.address1.strip().lower()


def _register_cache_entry(
    cache: Dict[str, Tuple[float, float]],
    canonical: Dict[str, Tuple[float, float]],
    address: str,
    lat: float,
    lon: float,
) -> None:
    canonical[address] = (lat, lon)
    normalized = address.lower()
    cache[normalized] = (lat, lon)
    first_line = address.splitlines()[0].strip().lower()
    if first_line:
        cache[first_line] = (lat, lon)


def _load_address_cache(cache_path: Path) -> Tuple[
    Dict[str, Tuple[float, float]],
    Dict[str, Tuple[float, float]],
    int,
    bool,
]:
    cache: Dict[str, Tuple[float, float]] = {}
    canonical: Dict[str, Tuple[float, float]] = {}
    if not cache_path.exists():
        return cache, canonical, 0, True

    valid_rows = 0
    has_header = False

    with cache_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for idx, row in enumerate(reader):
            if not row:
                continue
            if idx == 0 and row[0].strip().lower() == "address":
                has_header = True
                continue
            address = (row[0] if len(row) > 0 else "").strip()
            lat = Route._parse_float(row[1] if len(row) > 1 else None)
            lon = Route._parse_float(row[2] if len(row) > 2 else None)
            if not address or lat is None or lon is None:
                continue
            valid_rows += 1
            _register_cache_entry(cache, canonical, address, lat, lon)

    return cache, canonical, valid_rows, has_header


def _rewrite_cache_file(cache_path: Path, canonical: Dict[str, Tuple[float, float]]) -> None:
    with cache_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["address", "lat", "lon"])
        for address in sorted(canonical):
            lat, lon = canonical[address]
            writer.writerow([address, f"{lat:.8f}", f"{lon:.8f}"])


def _append_cache_entry(cache_path: Path, address: str, lat: float, lon: float) -> None:
    with cache_path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([address, f"{lat:.8f}", f"{lon:.8f}"])
        handle.flush()




def _within_victoria(lat: float, lon: float) -> bool:
    return (
        VICTORIA_BOUNDING_BOX["min_lat"] <= lat <= VICTORIA_BOUNDING_BOX["max_lat"]
        and VICTORIA_BOUNDING_BOX["min_lon"] <= lon <= VICTORIA_BOUNDING_BOX["max_lon"]
    )


def _distance_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius * c


def _address_with_notes(order: Order, separator: str = "\\n") -> str:
    parts: List[str] = []
    if order.address1.strip():
        parts.append(order.address1.strip())
    if order.address2.strip():
        parts.append(order.address2.strip())
    return separator.join(parts)

def _format_float(value: Optional[float]) -> str:
    if value is None:
        return ""
    return f"{value:.6f}"



def run_workflow() -> None:
    """Adjust the paths below and run this file in PyCharm to execute the workflow."""
    external_csv = Path("data/real_addresses.csv")  # TODO: update to the latest external CSV
    routes_csv = Path("data/route_centers.csv")  # optional route metadata lookup

    output_dir = Path("output") / datetime.now().strftime(OUTPUT_TIMESTAMP_FORMAT)
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {output_dir}")

    working_csv = output_dir / "orders_working.csv"  # internal, full-data CSV output
    simple_xlsx = output_dir / "orders_simple.xlsx"  # simplified workbook for sharing
    kml_path = output_dir / "orders_route.kml"  # optional Google Earth export

    routes: Optional[Dict[str, Route]] = None
    if routes_csv.exists():
        try:
            routes = Route.load_from_csv(routes_csv)
        except OSError as exc:
            print(f"Unable to load routes CSV at {routes_csv}: {exc}")
    else:
        print(f"Route CSV not found at {routes_csv}, continuing without routes")

    try:
        book = OrderBook.from_external_csv(external_csv, routes)
    except OSError as exc:
        print(f"Unable to load external CSV at {external_csv}: {exc}")
        return

    cache_path = Path("data/address_cache.csv")
    try:
        geocode_orders(book.orders, cache_path)
    except (RuntimeError, ValueError) as exc:
        print(exc)
        return

    # if routes:
    #     try:
    #         book.assign_routes(routes, max_orders_per_route=MAX_ORDERS_PER_ROUTE)
    #         print(f"Assigned routes with a cap of {MAX_ORDERS_PER_ROUTE} orders per route.")
    #     except ValueError as exc:
    #         print(f"Unable to assign routes: {exc}")
    #         return
    # else:
    #     print("No routes provided; skipping route assignment.")

    try:
        book.to_working_csv(working_csv)
        print(f"Saved working CSV to {working_csv}")
    except OSError as exc:
        print(f"Unable to write working CSV: {exc}")

    try:
        book.to_simple_xlsx(simple_xlsx)
        print(f"Saved simplified XLSX to {simple_xlsx}")
    except (RuntimeError, OSError) as exc:
        print(f"Unable to write XLSX: {exc}")

    missing_coords = [order for order in book.orders if order.lat is None or order.lon is None]
    if missing_coords:
        print(f"Skipping KML export: missing coordinates for {len(missing_coords)} orders.")
    else:
        try:
            book.to_kml(kml_path, connect_points=False)
            print(f"Saved KML to {kml_path}")
        except OSError as exc:
            print(f"Unable to write KML: {exc}")

if __name__ == "__main__":
    run_workflow()



























